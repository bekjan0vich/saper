import customtkinter as ctk
import random
import json
import csv
import os
import webbrowser
import platform
from tkinter import messagebox as mb
from datetime import datetime
from openpyxl import Workbook
from fpdf import FPDF

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class Minesweeper:
    def open_file(filepath):
        if platform.system() == "Windows":
            os.startfile(filepath)
        elif platform.system() == "Darwin":
            os.system(f"open '{filepath}'")
        else:
            os.system(f"xdg-open '{filepath}'")

    def __init__(self, master):
        self.master = master
        self.master.title("SAPPER")
        self.master.geometry("1000x700")
        self.master.resizable(False, False)

        # Тилдер
        self.current_lang = "ky"  # ky, ru, en
        self.translations = self.load_translations()

        # Оюн параметрлери
        self.rows = 9
        self.cols = 9
        self.mines = 10
        self.player_name = "Player"
        self.difficulty = "easy"

        # Custom параметрлер
        self.custom_rows = 10
        self.custom_cols = 10
        self.custom_mines = 15

        # Статистика
        self.games_played = 0
        self.games_won = 0
        self.total_time = 0

        # High Scores
        self.high_scores = []
        self.load_high_scores()

        self.show_main_menu()

    def load_translations(self):
        return {
            "ky": {
                "title": "САПЕР",
                "play": "ОЙНО",
                "setting": "ОЮН ПАРАМЕТРЛЕРИ",
                "quit": "ЧЫГУУ",
                "game_over": "ОЮН БҮТТҮ!",
                "game_over_msg": "КЕЙИНКИ ЖОЛУ БАКЫТЫҢЫЗДЫ СЫНАП КӨРҮҢҮЗ!",
                "you_win": "СИЗ ЖЕНИҢИЗ!",
                "win_msg": "КУТТУКТАЙМЫН {}! 🎉\nУбакыт: {}",
                "settings": "ОЮН ПАРАМЕТРЛЕРИ",
                "high_scores": "ҮЗДҮК НАТЫЙЖЕЛЕР",
                "player_name": "ОЮНЧУНУН АТЫ:",
                "difficulty": "КЫЙЫНДЫК:",
                "beginner": "ЖАҢЫ БАШТООЧУ",
                "easy": "ЖЕНИЛ",
                "normal": "ОРТО",
                "hard": "КЫЙЫН",
                "expert": "ЭКСПЕРТ",
                "custom": "ӨЗГӨРТМӨ",
                "custom_params": "ӨЗГӨРТМӨ (катар/тилке/мина):",
                "rows": "катар",
                "cols": "тилке",
                "mines": "мина",
                "save": "САКТОО",
                "back": "АРТКА",
                "saved": "САКТАЛДЫ",
                "saved_msg": "Оюн параметрлери ийгиликтүү сакталды!",
                "rank": "Орун",
                "name": "Аты",
                "time": "Убакыт",
                "close": "ЖАБУУ",
                "export": "ЭКСПОРТ",
                "export_title": "Экспорттоо",
                "format": "Формат:",
                "type": "Түрү:",
                "standard": "Стандарттуу",
                "detailed": "Толук",
                "summary": "Кыскача",
                "avg_time": "Орт. убакыт:",
                "win_percent": "Жеңүү %:",
                "flags_placed": "Желектер:",
                "level": "Денгээл",
                "analytics": "Аналитика",
                "preview": "Көрүү",
                "download": "Жүктөө",
                "reset_stats": "Статистиканы баштапкылоо",
                "add_custom": "Кошуу",
                "remove_custom": "Өчүрүү",
                "post": "POST",
                "usb": "USB",
                "gnd": "GND"
            },
            "ru": {
                "title": "САПЁР",
                "play": "ИГРАТЬ",
                "setting": "НАСТРОЙКИ",
                "quit": "ВЫХОД",
                "game_over": "ИГРА ОКОНЧЕНА!",
                "game_over_msg": "НЕ ТЕРЯЙТЕ НАДЕЖДУ!",
                "you_win": "ВЫ ПОБЕДИЛИ!",
                "win_msg": "ПОЗДРАВЛЯЕМ {}! 🎉\nВремя: {}",
                "settings": "НАСТРОЙКИ",
                "high_scores": "ЛУЧШИЕ РЕЗУЛЬТАТЫ",
                "player_name": "ИМЯ ИГРОКА:",
                "difficulty": "СЛОЖНОСТЬ:",
                "beginner": "НАЧИНАЮЩИЙ",
                "easy": "ЛЁГКИЙ",
                "normal": "СРЕДНИЙ",
                "hard": "СЛОЖНЫЙ",
                "expert": "ЭКСПЕРТ",
                "custom": "ПОЛЬЗОВАТЕЛЬСКАЯ",
                "custom_params": "Пользовательские (строки/колонки/мины):",
                "rows": "строки",
                "cols": "колонки",
                "mines": "мины",
                "save": "СОХРАНИТЬ",
                "back": "НАЗАД",
                "saved": "СОХРАНЕНО",
                "saved_msg": "Настройки успешно сохранены!",
                "rank": "Место",
                "name": "Имя",
                "time": "Время",
                "close": "ЗАКРЫТЬ",
                "export": "ЭКСПОРТ",
                "export_title": "Экспорт данных",
                "format": "Формат:",
                "type": "Тип:",
                "standard": "Стандартный",
                "detailed": "Детальный",
                "summary": "Краткий",
                "avg_time": "Ср. время:",
                "win_percent": "Побед %:",
                "flags_placed": "Флажков:",
                "level": "Уровень",
                "analytics": "Аналитика",
                "preview": "Просмотр",
                "download": "Скачать",
                "reset_stats": "Сбросить статистику",
                "add_custom": "Добавить",
                "remove_custom": "Удалить",
                "post": "POST",
                "usb": "USB",
                "gnd": "GND"
            },
            "en": {
                "title": "SAPPER",
                "play": "PLAY",
                "setting": "SETTINGS",
                "quit": "QUIT",
                "game_over": "GAME OVER!",
                "game_over_msg": "DON'T LOSE HOPE!",
                "you_win": "YOU WIN!",
                "win_msg": "Congratulations {}! 🎉\nTime: {}",
                "settings": "SETTINGS",
                "high_scores": "HIGH SCORES",
                "player_name": "PLAYER NAME:",
                "difficulty": "DIFFICULTY:",
                "beginner": "BEGINNER",
                "easy": "EASY",
                "normal": "NORMAL",
                "hard": "HARD",
                "expert": "EXPERT",
                "custom": "CUSTOM",
                "custom_params": "Custom (rows/cols/mines):",
                "rows": "rows",
                "cols": "cols",
                "mines": "mines",
                "save": "SAVE",
                "back": "BACK",
                "saved": "SAVED",
                "saved_msg": "Settings saved successfully!",
                "rank": "Rank",
                "name": "Name",
                "time": "Time",
                "close": "CLOSE",
                "export": "EXPORT",
                "export_title": "Export Data",
                "format": "Format:",
                "type": "Type:",
                "standard": "Standard",
                "detailed": "Detailed",
                "summary": "Summary",
                "avg_time": "Avg. Time:",
                "win_percent": "Win %:",
                "flags_placed": "Flags:",
                "level": "Level",
                "analytics": "Analytics",
                "preview": "Preview",
                "download": "Download",
                "reset_stats": "Reset Stats",
                "add_custom": "Add",
                "remove_custom": "Remove",
                "post": "POST",
                "usb": "USB",
                "gnd": "GND"
            }
        }

    def t(self, key):
        return self.translations[self.current_lang].get(key, key)

    def change_language(self, lang):
        self.current_lang = lang
        self.show_main_menu()

    def load_high_scores(self):
        try:
            with open("high_scores.json", "r") as f:
                self.high_scores = json.load(f)
        except:
            self.high_scores = []

    def save_high_scores(self):
        with open("high_scores.json", "w") as f:
            json.dump(self.high_scores, f)

    def clear_window(self):
        for widget in self.master.winfo_children():
            widget.destroy()

    def show_main_menu(self):
        self.clear_window()

        # Тил тандоо баскычтары
        lang_frame = ctk.CTkFrame(self.master)
        lang_frame.pack(pady=10)

        ctk.CTkButton(lang_frame, text="🇰🇬 Кыргызча", width=100,
                      command=lambda: self.change_language("ky")).pack(side="left", padx=5)
        ctk.CTkButton(lang_frame, text="🇷🇺 Русский", width=100,
                      command=lambda: self.change_language("ru")).pack(side="left", padx=5)
        ctk.CTkButton(lang_frame, text="🇬🇧 English", width=100,
                      command=lambda: self.change_language("en")).pack(side="left", padx=5)

        title = ctk.CTkLabel(self.master, text=self.t("title"), font=("Arial", 64, "bold"), text_color="red")
        title.pack(pady=60)

        play_btn = ctk.CTkButton(self.master, text=self.t("play"), width=250, height=60, font=("Arial", 24),
                                 command=self.start_game)
        play_btn.pack(pady=15)

        setting_btn = ctk.CTkButton(self.master, text=self.t("setting"), width=250, height=60, font=("Arial", 24),
                                    command=self.show_settings)
        setting_btn.pack(pady=15)

        out_btn = ctk.CTkButton(self.master, text=self.t("quit"), width=250, height=60, font=("Arial", 24),
                                fg_color="red", command=self.master.quit)
        out_btn.pack(pady=15)

    def start_game(self):
        if self.difficulty == "beginner":
            self.rows, self.cols, self.mines = 8, 8, 8
        elif self.difficulty == "easy":
            self.rows, self.cols, self.mines = 9, 9, 10
        elif self.difficulty == "normal":
            self.rows, self.cols, self.mines = 12, 12, 30
        elif self.difficulty == "hard":
            self.rows, self.cols, self.mines = 16, 16, 50
        elif self.difficulty == "expert":
            self.rows, self.cols, self.mines = 18, 18, 80
        elif self.difficulty == "custom":
            self.rows = self.custom_rows
            self.cols = self.custom_cols
            self.mines = self.custom_mines

        self.game = GameLogic(self.rows, self.cols, self.mines)
        self.show_game_screen()

    def show_game_screen(self):
        self.clear_window()

        # Тил тандоо (оюн ичинде да)
        lang_frame = ctk.CTkFrame(self.master)
        lang_frame.pack(pady=5)

        ctk.CTkButton(lang_frame, text="🇰🇬", width=40, command=lambda: self.change_language("ky")).pack(side="left", padx=2)
        ctk.CTkButton(lang_frame, text="🇷🇺", width=40, command=lambda: self.change_language("ru")).pack(side="left", padx=2)
        ctk.CTkButton(lang_frame, text="🇬🇧", width=40, command=lambda: self.change_language("en")).pack(side="left", padx=2)

        self.master.title(f"SAPER - {self.t('level')}: {self.t(self.difficulty)}")

        top_frame = ctk.CTkFrame(self.master)
        top_frame.pack(pady=10, fill="x")

        level_label = ctk.CTkLabel(top_frame, text=f"SAPER    {self.t('level')}: {self.t(self.difficulty)}",
                                   font=("Arial", 24, "bold"), text_color="red")
        level_label.pack(side="left", padx=20)

        main_container = ctk.CTkFrame(self.master)
        main_container.pack(pady=10, fill="both", expand=True)

        game_frame = ctk.CTkFrame(main_container)
        game_frame.pack(side="left", padx=20, pady=10)

        self.buttons = []
        for r in range(self.rows):
            row = []
            for c in range(self.cols):
                btn = ctk.CTkButton(game_frame, text="", width=45, height=45, corner_radius=5,
                                    font=("Arial", 16), fg_color="#3a3a3a",
                                    command=lambda r=r, c=c: self.game_click(r, c))
                btn.grid(row=r, column=c, padx=1, pady=1)
                btn.bind("<Button-3>", lambda e, r=r, c=c: self.flag_click(r, c))
                row.append(btn)
            self.buttons.append(row)

        analytics_frame = ctk.CTkFrame(main_container, width=280)
        analytics_frame.pack(side="right", padx=20, pady=10, fill="both")

        analytics_title = ctk.CTkLabel(analytics_frame, text=self.t("analytics"), font=("Arial", 24, "bold"))
        analytics_title.pack(pady=15)

        self.avg_time_label = ctk.CTkLabel(analytics_frame, text=self.t("avg_time"), font=("Arial", 16))
        self.avg_time_label.pack(pady=5)
        self.avg_time_value = ctk.CTkLabel(analytics_frame, text="0:00", font=("Arial", 20, "bold"))
        self.avg_time_value.pack(pady=5)

        self.win_percent_label = ctk.CTkLabel(analytics_frame, text=self.t("win_percent"), font=("Arial", 16))
        self.win_percent_label.pack(pady=5)
        self.win_percent_value = ctk.CTkLabel(analytics_frame, text="0%", font=("Arial", 20, "bold"),
                                              text_color="green")
        self.win_percent_value.pack(pady=5)

        self.flags_label = ctk.CTkLabel(analytics_frame, text=self.t("flags_placed"), font=("Arial", 16))
        self.flags_label.pack(pady=5)
        self.flags_value = ctk.CTkLabel(analytics_frame, text="0", font=("Arial", 20, "bold"), text_color="yellow")
        self.flags_value.pack(pady=5)

        export_btn = ctk.CTkButton(analytics_frame, text=self.t("export"), width=200, height=45,
                                   font=("Arial", 18), command=self.show_export_window)
        export_btn.pack(pady=30)

        # POST/USB/GND баскычтары
        post_frame = ctk.CTkFrame(analytics_frame)
        post_frame.pack(pady=20)
        ctk.CTkButton(post_frame, text=self.t("post"), width=80, fg_color="#333",
                      command=self.reset_stats).pack(side="left", padx=5)
        ctk.CTkButton(post_frame, text=self.t("usb"), width=80, fg_color="#333",
                      command=self.add_custom_difficulty).pack(side="left", padx=5)
        ctk.CTkButton(post_frame, text=self.t("gnd"), width=80, fg_color="#333",
                      command=self.remove_last_custom).pack(side="left", padx=5)

        self.update_analytics()
        self.game_start_time = datetime.now()
        self.update_game_grid()

    def reset_stats(self):
        self.games_played = 0
        self.games_won = 0
        self.total_time = 0
        self.update_analytics()
        mb.showinfo("Stats", "Statistics reset successfully!")

    def add_custom_difficulty(self):
        # Кошумча кыйындык деңгээлин кошуу
        self.difficulties.append(f"custom_{len(self.difficulties)}")
        mb.showinfo("Custom", "New difficulty level added!")

    def remove_last_custom(self):
        # Акыркы кошулганды өчүрүү
        if len(self.difficulties) > 6:
            self.difficulties.pop()
            mb.showinfo("Removed", "Last custom difficulty removed!")

    def update_analytics(self):
        if self.games_played > 0:
            avg_sec = self.total_time / self.games_played
            minutes = int(avg_sec // 60)
            seconds = int(avg_sec % 60)
            self.avg_time_value.configure(text=f"{minutes}:{seconds:02d}")
            win_rate = (self.games_won / self.games_played) * 100
            self.win_percent_value.configure(text=f"{int(win_rate)}%")

    def update_game_grid(self):
        for r in range(self.rows):
            for c in range(self.cols):
                if self.game.revealed[r][c]:
                    if self.game.board[r][c] == -1:
                        self.buttons[r][c].configure(text="💣", fg_color="#8B0000")
                    elif self.game.board[r][c] == 0:
                        self.buttons[r][c].configure(text="", fg_color="#555555")
                    else:
                        colors = ["", "blue", "green", "red", "darkblue", "brown", "cyan", "black", "gray"]
                        self.buttons[r][c].configure(text=str(self.game.board[r][c]),
                                                     fg_color="#555555", text_color=colors[self.game.board[r][c]])
                elif self.game.flags[r][c]:
                    self.buttons[r][c].configure(text="🚩", fg_color="#2d5a2d")
                else:
                    self.buttons[r][c].configure(text="", fg_color="#3a3a3a")

        self.flags_value.configure(text=str(self.game.flags_placed))
        self.master.update()

    def game_click(self, r, c):
        if self.game.game_over:
            return
        if self.game.flags[r][c]:
            return

        if self.game.board[r][c] == -1:
            self.game.game_over = True
            for rr in range(self.rows):
                for cc in range(self.cols):
                    if self.game.board[rr][cc] == -1:
                        self.buttons[rr][cc].configure(text="💣", fg_color="#8B0000")
            mb.showinfo(self.t("game_over"), self.t("game_over_msg"))
            self.games_played += 1
            self.update_analytics()
            self.show_main_menu()
            return

        self.reveal_cells(r, c)
        self.update_game_grid()

        revealed_count = sum(sum(row) for row in self.game.revealed)
        if revealed_count == (self.rows * self.cols) - self.mines:
            self.game.game_over = True
            self.games_played += 1
            self.games_won += 1
            play_time = (datetime.now() - self.game_start_time).total_seconds()
            self.total_time += play_time

            minutes = int(play_time // 60)
            seconds = int(play_time % 60)
            time_str = f"{minutes}:{seconds:02d}"
            self.high_scores.append({"name": self.player_name, "time": time_str, "seconds": play_time})
            self.high_scores.sort(key=lambda x: x["seconds"])
            self.high_scores = self.high_scores[:6]
            self.save_high_scores()

            mb.showinfo(self.t("you_win"), self.t("win_msg").format(self.player_name, time_str))
            self.update_analytics()
            self.show_main_menu()

    def reveal_cells(self, r, c):
        if self.game.revealed[r][c] or self.game.flags[r][c]:
            return
        self.game.revealed[r][c] = True
        if self.game.board[r][c] == 0:
            for dr in [-1, 0, 1]:
                for dc in [-1, 0, 1]:
                    nr, nc = r + dr, c + dc
                    if 0 <= nr < self.rows and 0 <= nc < self.cols:
                        if not self.game.revealed[nr][nc] and not self.game.flags[nr][nc]:
                            self.reveal_cells(nr, nc)

    def flag_click(self, r, c):
        if self.game.game_over or self.game.revealed[r][c]:
            return
        self.game.flags[r][c] = not self.game.flags[r][c]
        if self.game.flags[r][c]:
            self.game.flags_placed += 1
        else:
            self.game.flags_placed -= 1
        self.update_game_grid()

    def show_settings(self):
        self.clear_window()

        # Тил тандоо
        lang_frame = ctk.CTkFrame(self.master)
        lang_frame.pack(pady=10)
        ctk.CTkButton(lang_frame, text="🇰🇬", width=40, command=lambda: self.change_language("ky")).pack(side="left", padx=2)
        ctk.CTkButton(lang_frame, text="🇷🇺", width=40, command=lambda: self.change_language("ru")).pack(side="left", padx=2)
        ctk.CTkButton(lang_frame, text="🇬🇧", width=40, command=lambda: self.change_language("en")).pack(side="left", padx=2)

        ctk.CTkLabel(self.master, text=self.t("settings"), font=("Arial", 36, "bold")).pack(pady=20)

        hs_btn = ctk.CTkButton(self.master, text=self.t("high_scores"), width=200, height=40,
                               font=("Arial", 18), command=self.show_high_scores)
        hs_btn.pack(pady=10)

        ctk.CTkLabel(self.master, text=self.t("player_name"), font=("Arial", 16)).pack(pady=10)
        self.name_entry = ctk.CTkEntry(self.master, width=250, placeholder_text="Enter name")
        self.name_entry.insert(0, self.player_name)
        self.name_entry.pack(pady=5)

        ctk.CTkLabel(self.master, text=self.t("difficulty"), font=("Arial", 16)).pack(pady=10)
        self.difficulty_var = ctk.StringVar(value=self.difficulty)
        difficulty_frame = ctk.CTkFrame(self.master)
        difficulty_frame.pack(pady=5)

        levels = ["beginner", "easy", "normal", "hard", "expert", "custom"]
        level_names = [self.t("beginner"), self.t("easy"), self.t("normal"),
                       self.t("hard"), self.t("expert"), self.t("custom")]
        for level, name in zip(levels, level_names):
            rb = ctk.CTkRadioButton(difficulty_frame, text=name, variable=self.difficulty_var, value=level)
            rb.pack(side="left", padx=8)

        custom_frame = ctk.CTkFrame(self.master)
        custom_frame.pack(pady=10)

        ctk.CTkLabel(custom_frame, text=self.t("custom_params"), font=("Arial", 14)).grid(row=0, column=0, padx=5)
        self.custom_rows_entry = ctk.CTkEntry(custom_frame, width=60, placeholder_text=self.t("rows"))
        self.custom_rows_entry.insert(0, str(self.custom_rows))
        self.custom_rows_entry.grid(row=0, column=1, padx=5)

        self.custom_cols_entry = ctk.CTkEntry(custom_frame, width=60, placeholder_text=self.t("cols"))
        self.custom_cols_entry.insert(0, str(self.custom_cols))
        self.custom_cols_entry.grid(row=0, column=2, padx=5)

        self.custom_mines_entry = ctk.CTkEntry(custom_frame, width=60, placeholder_text=self.t("mines"))
        self.custom_mines_entry.insert(0, str(self.custom_mines))
        self.custom_mines_entry.grid(row=0, column=3, padx=5)

        save_btn = ctk.CTkButton(self.master, text=self.t("save"), width=200, height=45,
                                 font=("Arial", 18), fg_color="green", command=self.save_settings)
        save_btn.pack(pady=20)

        back_btn = ctk.CTkButton(self.master, text=self.t("back"), width=200, height=40,
                                 command=self.show_main_menu)
        back_btn.pack(pady=10)

    def save_settings(self):
        self.player_name = self.name_entry.get()
        self.difficulty = self.difficulty_var.get()

        try:
            self.custom_rows = int(self.custom_rows_entry.get())
            self.custom_cols = int(self.custom_cols_entry.get())
            self.custom_mines = int(self.custom_mines_entry.get())
        except:
            pass

        mb.showinfo(self.t("saved"), self.t("saved_msg"))
        self.show_main_menu()

    def show_high_scores(self):
        hs_window = ctk.CTkToplevel(self.master)
        hs_window.title(self.t("high_scores"))
        hs_window.geometry("500x400")
        hs_window.resizable(False, False)

        ctk.CTkLabel(hs_window, text=self.t("high_scores"), font=("Arial", 28, "bold")).pack(pady=20)

        frame = ctk.CTkFrame(hs_window)
        frame.pack(pady=10, padx=20, fill="both", expand=True)

        ctk.CTkLabel(frame, text=self.t("rank"), width=80, font=("Arial", 16, "bold")).grid(row=0, column=0, padx=10, pady=5)
        ctk.CTkLabel(frame, text=self.t("name"), width=150, font=("Arial", 16, "bold")).grid(row=0, column=1, padx=10, pady=5)
        ctk.CTkLabel(frame, text=self.t("time"), width=100, font=("Arial", 16, "bold")).grid(row=0, column=2, padx=10, pady=5)

        for i, score in enumerate(self.high_scores[:6]):
            ctk.CTkLabel(frame, text=f"{i + 1}", width=80, font=("Arial", 14)).grid(row=i + 1, column=0, padx=10, pady=3)
            ctk.CTkLabel(frame, text=score["name"], width=150, font=("Arial", 14)).grid(row=i + 1, column=1, padx=10, pady=3)
            ctk.CTkLabel(frame, text=score["time"], width=100, font=("Arial", 14)).grid(row=i + 1, column=2, padx=10, pady=3)

        for i in range(len(self.high_scores), 6):
            ctk.CTkLabel(frame, text=f"{i + 1}", width=80, font=("Arial", 14)).grid(row=i + 1, column=0, padx=10, pady=3)
            ctk.CTkLabel(frame, text="???", width=150, font=("Arial", 14)).grid(row=i + 1, column=1, padx=10, pady=3)
            ctk.CTkLabel(frame, text="0.00", width=100, font=("Arial", 14)).grid(row=i + 1, column=2, padx=10, pady=3)

        close_btn = ctk.CTkButton(hs_window, text=self.t("close"), width=150, command=hs_window.destroy)
        close_btn.pack(pady=20)

    def show_export_window(self):
        export_window = ctk.CTkToplevel(self.master)
        export_window.title(self.t("export_title"))
        export_window.geometry("500x600")
        export_window.resizable(False, False)

        ctk.CTkLabel(export_window, text=self.t("export_title"), font=("Arial", 28, "bold")).pack(pady=20)

        ctk.CTkLabel(export_window, text=self.t("format"), font=("Arial", 16)).pack(pady=10)
        format_var = ctk.StringVar(value="JSON")
        format_frame = ctk.CTkFrame(export_window)
        format_frame.pack(pady=5)
        for fmt in ["JSON", "CSV", "Excel", "PDF", "HTML"]:
            rb = ctk.CTkRadioButton(format_frame, text=fmt, variable=format_var, value=fmt)
            rb.pack(side="left", padx=8)

        ctk.CTkLabel(export_window, text=self.t("type"), font=("Arial", 16)).pack(pady=10)
        type_var = ctk.StringVar(value=self.t("standard"))
        type_frame = ctk.CTkFrame(export_window)
        type_frame.pack(pady=5)
        for t in [self.t("standard"), self.t("detailed"), self.t("summary")]:
            rb = ctk.CTkRadioButton(type_frame, text=t, variable=type_var, value=t)
            rb.pack(side="left", padx=8)

        def get_export_type(val):
            if val == self.t("standard"):
                return "Standard"
            elif val == self.t("detailed"):
                return "Detailed"
            return "Summary"

        preview_btn = ctk.CTkButton(export_window, text=self.t("preview"), width=200, height=40,
                                    font=("Arial", 16),
                                    command=lambda: self.preview_export(format_var.get(), get_export_type(type_var.get())))
        preview_btn.pack(pady=20)

        download_btn = ctk.CTkButton(export_window, text=self.t("download"), width=200, height=40,
                                     font=("Arial", 16), fg_color="green",
                                     command=lambda: self.download_export(format_var.get(), get_export_type(type_var.get())))
        download_btn.pack(pady=10)

    def preview_export(self, format_type, export_type):
        data = self.get_export_data(export_type)
        if format_type == "JSON":
            preview_text = json.dumps(data, indent=2)[:500]
        elif format_type == "CSV":
            preview_text = "CSV preview:\n" + "\n".join([",".join(map(str, row)) for row in data[:5]])
        else:
            preview_text = str(data)[:500]
        mb.showinfo(self.t("preview"), preview_text)

    def download_export(self, format_type, export_type):
        data = self.get_export_data(export_type)
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        if format_type == "JSON":
            filename += ".json"
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            mb.showinfo(self.t("download"), f"Exported to {filename}")

        elif format_type == "CSV":
            filename += ".csv"
            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                if isinstance(data[0], list):
                    writer.writerows(data)
                else:
                    for key, val in data.items():
                        writer.writerow([key, val])
            mb.showinfo(self.t("download"), f"Exported to {filename}")

        elif format_type == "Excel":
            filename += ".xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sapper Export"
            if isinstance(data[0], list):
                for row in data:
                    ws.append(row)
            else:
                for key, val in data.items():
                    ws.append([key, val])
            wb.save(filename)
            mb.showinfo(self.t("download"), f"Exported to {filename}")

        elif format_type == "PDF":
            filename += ".pdf"
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt="Sapper Game Export", ln=1, align="C")
            pdf.cell(200, 10, txt=f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1, align="C")
            pdf.ln(10)

            if isinstance(data[0], list):
                for row in data:
                    pdf.cell(200, 8, txt=", ".join(str(x) for x in row), ln=1)
            else:
                for key, val in data.items():
                    pdf.cell(200, 8, txt=f"{key}: {val}", ln=1)
            pdf.output(filename)
            mb.showinfo(self.t("download"), f"Exported to {filename}")

        elif format_type == "HTML":
            filename += ".html"
            html_content = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>Sapper Export</title>
<style>
body {{ font-family: Arial; background: #1e1e1e; color: white; padding: 20px; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ border: 1px solid #555; padding: 8px; text-align: left; }}
th {{ background: #333; }}
</style>
</head>
<body>
<h1>Sapper Game Export</h1>
<p>Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
<p>Player: {self.player_name}</p>
<p>Difficulty: {self.difficulty}</p>
<h2>Game Data</h2>
<table>
"""
            if isinstance(data[0], list):
                for row in data:
                    html_content += "<tr>"
                    for cell in row:
                        html_content += f"<td>{cell}</td>"
                    html_content += "</tr>"
            else:
                for key, val in data.items():
                    html_content += f"<tr><td>{key}</td><td>{val}</td></tr>"

            html_content += """</table></body></html>"""
            with open(filename, "w", encoding="utf-8") as f:
                f.write(html_content)
            webbrowser.open(filename)
            mb.showinfo(self.t("download"), f"Exported to {filename} (opened in browser)")

    def get_export_data(self, export_type):
        if export_type == "Standard":
            return {
                "Player": self.player_name,
                "Difficulty": self.difficulty,
                "Mines": self.mines,
                "Flags Placed": self.game.flags_placed,
                "Status": "Win" if self.game.game_over and not self.game.game_over else "Playing"
            }
        elif export_type == "Detailed":
            board_data = []
            for r in range(self.rows):
                row = []
                for c in range(self.cols):
                    if self.game.revealed[r][c]:
                        row.append(self.game.board[r][c])
                    elif self.game.flags[r][c]:
                        row.append("F")
                    else:
                        row.append("?")
                board_data.append(row)
            return board_data
        else:
            return {
                "Player": self.player_name,
                "Difficulty": self.difficulty,
                "Rows": self.rows,
                "Cols": self.cols,
                "Mines": self.mines,
                "Flags Placed": self.game.flags_placed,
                "Game Over": self.game.game_over,
                "Export Time": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }


class GameLogic:
    def __init__(self, rows, cols, mines):
        self.rows = rows
        self.cols = cols
        self.mines = mines
        self.board = []
        self.revealed = []
        self.flags = []
        self.game_over = False
        self.flags_placed = 0
        self.create_board()

    def create_board(self):
        self.board = [[0 for _ in range(self.cols)] for _ in range(self.rows)]
        self.revealed = [[False for _ in range(self.cols)] for _ in range(self.rows)]
        self.flags = [[False for _ in range(self.cols)] for _ in range(self.rows)]
        self.flags_placed = 0
        self.game_over = False

        mines_placed = 0
        while mines_placed < self.mines:
            r = random.randint(0, self.rows - 1)
            c = random.randint(0, self.cols - 1)
            if self.board[r][c] != -1:
                self.board[r][c] = -1
                mines_placed += 1

        for r in range(self.rows):
            for c in range(self.cols):
                if self.board[r][c] == -1:
                    continue
                count = 0
                for dr in [-1, 0, 1]:
                    for dc in [-1, 0, 1]:
                        nr, nc = r + dr, c + dc
                        if 0 <= nr < self.rows and 0 <= nc < self.cols and self.board[nr][nc] == -1:
                            count += 1
                self.board[r][c] = count


if __name__ == "__main__":
    root = ctk.CTk()
    app = Minesweeper(root)
    root.mainloop()